from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count, F
from django.utils import timezone
from .models import Product, Category, Supplier, StockMovement, Laboratory


# في stock/views.py
# استبدل دالة stock_dashboard بهذا الكود المحسّن

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, F, Q
from decimal import Decimal

from .models import Product, Category, Supplier, StockMovement, Laboratory

# في ملف stock/views.py
# استبدل دالة dashboard بهذا الكود المصلح:

@login_required
def dashboard(request):
    """لوحة المخزون مع الإحصائيات والحركات الفعلية"""
    
    # ==========================================
    # الإحصائيات الأساسية
    # ==========================================
    
    # جميع المنتجات النشطة
    active_products = Product.objects.filter(is_active=True)
    
    # إجمالي عدد المنتجات
    total_products = active_products.count()
    
    # المنتجات منخفضة المخزون
    low_stock_products = active_products.filter(
        quantity__lte=F('min_quantity'),
        quantity__gt=0
    ).order_by('quantity')
    low_stock_count = low_stock_products.count()
    
    # المنتجات النافذة
    out_of_stock = active_products.filter(quantity=0).count()
    
    # قيمة المخزون الإجمالية (بسعر البيع)
    total_value = active_products.aggregate(
        total=Sum(F('quantity') * F('selling_price'))
    )['total'] or Decimal('0')
    
    # ==========================================
    # إحصائيات التصنيفات - تم إصلاح التعارض
    # ==========================================
    categories_raw = Category.objects.filter(is_active=True).annotate(
        products_count_value=Count('products', filter=Q(products__is_active=True)),
        total_quantity_value=Sum('products__quantity', filter=Q(products__is_active=True))
    ).order_by('-products_count_value')[:5]
    
    # تحويل النتائج إلى قائمة من القواميس لتجنب التعارض
    categories_stats = []
    for cat in categories_raw:
        categories_stats.append({
            'id': cat.id,
            'name': cat.name,
            'icon': cat.icon,
            'color': cat.color,
            'products_count': cat.products_count_value,
            'total_quantity': cat.total_quantity_value or 0,
        })
    
    # ==========================================
    # آخر حركات المخزون
    # ==========================================
    recent_movements = StockMovement.objects.select_related(
        'product', 
        'product__category'
    ).order_by('-created_at')[:10]
    
    # ==========================================
    # الموردين النشطين
    # ==========================================
    active_suppliers = Supplier.objects.filter(is_active=True).count()
    
    # ==========================================
    # المعامل النشطة
    # ==========================================
    active_laboratories = Laboratory.objects.filter(is_active=True).count()
    
    # ==========================================
    # السياق (Context)
    # ==========================================
    context = {
        # الإحصائيات
        'total_products': total_products,
        'low_stock_count': low_stock_count,
        'out_of_stock': out_of_stock,
        'total_value': total_value,
        'active_suppliers': active_suppliers,
        'active_laboratories': active_laboratories,
        
        # المنتجات منخفضة المخزون (للعرض)
        'low_stock_products': low_stock_products[:5],
        
        # التصنيفات (كقواميس بدون تعارض)
        'categories_stats': categories_stats,
        
        # الحركات الفعلية
        'recent_movements': recent_movements,
    }
    
    return render(request, 'stock/dashboard.html', context)
@login_required
def products_list(request):
    """قائمة المنتجات"""
    
    search_query = request.GET.get('search', '')
    category_id = request.GET.get('category', '')
    status = request.GET.get('status', '')
    
    products = Product.objects.select_related('category', 'supplier').filter(is_active=True)
    
    if search_query:
        products = products.filter(
            Q(item_name__icontains=search_query) |
            Q(barcode__icontains=search_query)
        )
    
    if category_id:
        products = products.filter(category_id=category_id)
    
    if status == 'low':
        products = products.filter(quantity__lte=F('min_quantity'), quantity__gt=0)
    elif status == 'out':
        products = products.filter(quantity=0)
    
    products = products.order_by('-created_at')
    
    categories = Category.objects.filter(is_active=True)
    suppliers = Supplier.objects.filter(is_active=True)  # ← أضيفي هذا السطر
    
    context = {
        'products': products,
        'categories': categories,
        'suppliers': suppliers,  # ← أضيفي هذا السطر
        'search_query': search_query,
        'category_id': category_id,
        'status': status,
    }
    
    return render(request, 'stock/products_list.html', context)


@login_required
def product_detail(request, pk):
    """تفاصيل المنتج"""
    
    product = get_object_or_404(Product, pk=pk)
    movements = product.movements.all()[:20]
    
    context = {
        'product': product,
        'movements': movements,
    }
    
    return render(request, 'stock/product_detail.html', context)


@login_required
def categories_list(request):
    """قائمة التصنيفات"""
    
    categories = Category.objects.filter(is_active=True).order_by('name')
    
    # إضافة الإحصائيات يدوياً
    for category in categories:
        category.products_count_display = category.products.filter(is_active=True).count()
        category.total_quantity_display = category.products.filter(
            is_active=True
        ).aggregate(total=Sum('quantity'))['total'] or 0
    
    context = {
        'categories': categories,
    }
    
    return render(request, 'stock/categories_list.html', context)


@login_required
def suppliers_list(request):
    """قائمة الموردين"""
    
    suppliers = Supplier.objects.filter(is_active=True).order_by('-created_at')
    
    # إضافة عدد المنتجات يدوياً
    for supplier in suppliers:
        supplier.products_count_display = supplier.products.filter(is_active=True).count()
    
    context = {
        'suppliers': suppliers,
    }
    
    return render(request, 'stock/suppliers_list.html', context)


@login_required
def laboratory(request):
    """صفحة المختبر"""
    
    return render(request, 'stock/laboratory.html')


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Category, Supplier, Product

# Category Views
def category_add(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        icon = request.POST.get('icon', 'fas fa-cube')
        color = request.POST.get('color', '#4A9EAD')
        
        Category.objects.create(
            name=name,
            description=description,
            icon=icon,
            color=color
        )
        
        messages.success(request, f'تم إضافة التصنيف "{name}" بنجاح')
        return redirect('stock:categories')
    
    return redirect('stock:categories')

def category_edit(request, pk):
    category = get_object_or_404(Category, pk=pk)
    
    if request.method == 'POST':
        category.name = request.POST.get('name')
        category.description = request.POST.get('description', '')
        category.icon = request.POST.get('icon', 'fas fa-cube')
        category.color = request.POST.get('color', '#4A9EAD')
        category.save()
        
        messages.success(request, f'تم تعديل التصنيف "{category.name}" بنجاح')
        return redirect('stock:categories')
    
    return redirect('stock:categories')

def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    name = category.name
    category.delete()
    
    messages.success(request, f'تم حذف التصنيف "{name}" بنجاح')
    return redirect('stock:categories')

# Supplier Views
def supplier_add(request):
    if request.method == 'POST':
        company_name = request.POST.get('company_name')
        phone = request.POST.get('phone')
        email = request.POST.get('email', '')
        address = request.POST.get('address', '')
        representative_name = request.POST.get('representative_name', '')
        representative_phone = request.POST.get('representative_phone', '')
        notes = request.POST.get('notes', '')
        
        Supplier.objects.create(
            company_name=company_name,
            phone=phone,
            email=email,
            address=address,
            representative_name=representative_name,
            representative_phone=representative_phone,
            notes=notes
        )
        
        messages.success(request, f'تم إضافة المورد "{company_name}" بنجاح')
        return redirect('stock:suppliers')
    
    return redirect('stock:suppliers')

def supplier_edit(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    
    if request.method == 'POST':
        supplier.company_name = request.POST.get('company_name')
        supplier.phone = request.POST.get('phone')
        supplier.email = request.POST.get('email', '')
        supplier.address = request.POST.get('address', '')
        supplier.representative_name = request.POST.get('representative_name', '')
        supplier.representative_phone = request.POST.get('representative_phone', '')
        supplier.notes = request.POST.get('notes', '')
        supplier.save()
        
        messages.success(request, f'تم تعديل المورد "{supplier.company_name}" بنجاح')
        return redirect('stock:suppliers')
    
    return redirect('stock:suppliers')

def supplier_delete(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    name = supplier.company_name
    supplier.delete()
    
    messages.success(request, f'تم حذف المورد "{name}" بنجاح')
    return redirect('stock:suppliers')

# Product Views
def product_add(request):
    if request.method == 'POST':
        item_name = request.POST.get('item_name')
        barcode = request.POST.get('barcode')
        category_id = request.POST.get('category')
        supplier_id = request.POST.get('supplier')
        quantity = int(request.POST.get('quantity', 0))
        min_quantity = int(request.POST.get('min_quantity', 5))
        box_number = request.POST.get('box_number', '')
        cost_price = float(request.POST.get('cost_price', 0))
        selling_price = float(request.POST.get('selling_price', 0))
        description = request.POST.get('description', '')
        notes = request.POST.get('notes', '')
        
        product = Product.objects.create(
            item_name=item_name,
            barcode=barcode,
            quantity=quantity,
            min_quantity=min_quantity,
            box_number=box_number,
            cost_price=cost_price,
            selling_price=selling_price,
            description=description,
            notes=notes
        )
        
        if category_id:
            product.category_id = category_id
        if supplier_id:
            product.supplier_id = supplier_id
        product.save()
        
        messages.success(request, f'تم إضافة المنتج "{item_name}" بنجاح')
        return redirect('stock:products')
    
    return redirect('stock:products')

def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    
    if request.method == 'POST':
        product.item_name = request.POST.get('item_name')
        product.barcode = request.POST.get('barcode')
        category_id = request.POST.get('category')
        supplier_id = request.POST.get('supplier')
        product.quantity = int(request.POST.get('quantity', 0))
        product.min_quantity = int(request.POST.get('min_quantity', 5))
        product.box_number = request.POST.get('box_number', '')
        product.cost_price = float(request.POST.get('cost_price', 0))
        product.selling_price = float(request.POST.get('selling_price', 0))
        product.description = request.POST.get('description', '')
        product.notes = request.POST.get('notes', '')
        
        if category_id:
            product.category_id = category_id
        else:
            product.category = None
            
        if supplier_id:
            product.supplier_id = supplier_id
        else:
            product.supplier = None
            
        product.save()
        
        messages.success(request, f'تم تعديل المنتج "{product.item_name}" بنجاح')
        return redirect('stock:products')
    
    return redirect('stock:products')

def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    name = product.item_name
    product.delete()
    
    messages.success(request, f'تم حذف المنتج "{name}" بنجاح')
    return redirect('stock:products')


from django.http import JsonResponse

@login_required
def supplier_detail_json(request, pk):
    """إرجاع تفاصيل المورد كـ JSON"""
    supplier = get_object_or_404(Supplier, pk=pk)
    
    data = {
        'id': supplier.id,
        'company_name': supplier.company_name,
        'supplier_code': supplier.supplier_code,
        'phone': supplier.phone,
        'email': supplier.email or '',
        'address': supplier.address or '',
        'representative_name': supplier.representative_name or '',
        'representative_phone': supplier.representative_phone or '',
        'notes': supplier.notes or '',
        'is_active': supplier.is_active,
        'products_count': supplier.products.filter(is_active=True).count(),
        'created_at': supplier.created_at.strftime('%Y-%m-%d %H:%M'),
        'updated_at': supplier.updated_at.strftime('%Y-%m-%d %H:%M'),
    }
    
    return JsonResponse(data)




import pandas as pd
import openpyxl
from openpyxl import Workbook
from django.http import HttpResponse
from io import BytesIO

@login_required
def import_products(request):
    """صفحة استيراد المنتجات"""
    return render(request, 'stock/import_products.html')


@login_required
def download_template(request):
    """تحميل ملف Excel النموذجي"""
    
    # إنشاء workbook جديد
    wb = Workbook()
    ws = wb.active
    ws.title = "المنتجات"
    
    # Headers
    headers = [
        'item_name',
        'barcode',
        'quantity',
        'cost_price',
        'selling_price',
        'category',
        'supplier',
        'min_quantity',
        'box_number',
        'description',
        'notes'
    ]
    
    # Headers بالعربي (صف ثاني)
    headers_ar = [
        'اسم المنتج *',
        'الباركود *',
        'الكمية',
        'سعر التكلفة',
        'سعر البيع',
        'التصنيف',
        'المورد',
        'الحد الأدنى',
        'رقم الصندوق',
        'الوصف',
        'ملاحظات'
    ]
    
    # كتابة Headers الإنجليزية
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="4A9EAD", end_color="4A9EAD", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    # كتابة Headers العربية
    for col, header in enumerate(headers_ar, start=1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="E7F3F5", end_color="E7F3F5", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    # أمثلة
    examples = [
        ['نظارة شمسية راي بان', '1234567890', 50, 150.00, 250.00, 'نظارات شمسية', 'شركة النظارات', 10, 'BOX-001', 'نظارة شمسية عالية الجودة', ''],
        ['عدسات لاصقة يومية', '0987654321', 100, 80.00, 120.00, 'عدسات', 'شركة العدسات', 20, 'BOX-002', 'عدسات يومية مريحة', ''],
        ['نظارة طبية', '1122334455', 30, 200.00, 350.00, 'نظارات طبية', 'شركة النظارات', 5, 'BOX-003', '', 'موديل جديد'],
    ]
    
    for row_idx, example in enumerate(examples, start=3):
        for col_idx, value in enumerate(example, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # تعديل عرض الأعمدة
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # حفظ الملف
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # إرجاع الملف
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="products_template.xlsx"'
    
    return response


@login_required
def import_products_process(request):
    """معالجة ملف الاستيراد"""
    
    if request.method != 'POST':
        return redirect('stock:import_products')
    
    if 'file' not in request.FILES:
        messages.error(request, 'الرجاء اختيار ملف')
        return redirect('stock:import_products')
    
    file = request.FILES['file']
    skip_duplicates = request.POST.get('skip_duplicates') == 'on'
    update_existing = request.POST.get('update_existing') == 'on'
    
    # قراءة الملف
    try:
        if file.name.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
    except Exception as e:
        messages.error(request, f'خطأ في قراءة الملف: {str(e)}')
        return redirect('stock:import_products')
    
    # التحقق من الأعمدة المطلوبة
    required_columns = ['item_name', 'barcode']
    missing_columns = [col for col in required_columns if col not in df.columns]
    
    if missing_columns:
        messages.error(request, f'الأعمدة التالية مفقودة: {", ".join(missing_columns)}')
        return redirect('stock:import_products')
    
    # معالجة البيانات
    results = {
        'success': 0,
        'skipped': 0,
        'errors': 0,
        'error_details': []
    }
    
    for index, row in df.iterrows():
        try:
            # تخطي الصف إذا كان فارغاً
            if pd.isna(row.get('item_name')) or pd.isna(row.get('barcode')):
                continue
            
            barcode = str(row['barcode']).strip()
            
            # التحقق من وجود المنتج
            existing_product = Product.objects.filter(barcode=barcode).first()
            
            if existing_product:
                if skip_duplicates and not update_existing:
                    results['skipped'] += 1
                    continue
                elif update_existing:
                    product = existing_product
                else:
                    results['skipped'] += 1
                    continue
            else:
                product = Product()
            
            # ملء البيانات
            product.item_name = str(row['item_name']).strip()
            product.barcode = barcode
            product.quantity = int(row.get('quantity', 0)) if pd.notna(row.get('quantity')) else 0
            product.cost_price = float(row.get('cost_price', 0)) if pd.notna(row.get('cost_price')) else 0
            product.selling_price = float(row.get('selling_price', 0)) if pd.notna(row.get('selling_price')) else 0
            product.min_quantity = int(row.get('min_quantity', 5)) if pd.notna(row.get('min_quantity')) else 5
            product.box_number = str(row.get('box_number', '')) if pd.notna(row.get('box_number')) else ''
            product.description = str(row.get('description', '')) if pd.notna(row.get('description')) else ''
            product.notes = str(row.get('notes', '')) if pd.notna(row.get('notes')) else ''
            
            # التصنيف
            if pd.notna(row.get('category')):
                category_name = str(row['category']).strip()
                category, _ = Category.objects.get_or_create(
                    name=category_name,
                    defaults={'icon': 'fas fa-cube', 'color': '#4A9EAD'}
                )
                product.category = category
            
            # المورد
            if pd.notna(row.get('supplier')):
                supplier_name = str(row['supplier']).strip()
                supplier, _ = Supplier.objects.get_or_create(
                    company_name=supplier_name,
                    defaults={'phone': '500000000'}
                )
                product.supplier = supplier
            
            product.save()
            results['success'] += 1
            
        except Exception as e:
            results['errors'] += 1
            results['error_details'].append(f'صف {index + 2}: {str(e)}')
    
    # عرض النتائج
    if results['success'] > 0:
        messages.success(request, f'تم استيراد {results["success"]} منتج بنجاح')
    
    if results['skipped'] > 0:
        messages.warning(request, f'تم تخطي {results["skipped"]} منتج')
    
    if results['errors'] > 0:
        messages.error(request, f'حدثت {results["errors"]} أخطاء')
    
    context = {
        'results': results
    }
    
    return render(request, 'stock/import_products.html', context)




# وأيضاً عدّل دالة laboratory لإزالة products_count_display:

@login_required
def laboratory(request):
    """قائمة المعامل"""
    
    laboratories = Laboratory.objects.filter(is_active=True).order_by('-created_at')
    
    # لا حاجة لإضافة products_count_display لأن المعمل لا يحتوي على منتجات
    
    context = {
        'laboratories': laboratories,
    }
    
    return render(request, 'stock/laboratory.html', context)


@login_required
def laboratory_add(request):
    if request.method == 'POST':
        company_name = request.POST.get('company_name')
        phone = request.POST.get('phone')
        email = request.POST.get('email', '')
        address = request.POST.get('address', '')
        representative_name = request.POST.get('representative_name', '')
        representative_phone = request.POST.get('representative_phone', '')
        notes = request.POST.get('notes', '')
        
        Laboratory.objects.create(
            company_name=company_name,
            phone=phone,
            email=email,
            address=address,
            representative_name=representative_name,
            representative_phone=representative_phone,
            notes=notes
        )
        
        messages.success(request, f'تم إضافة المعمل "{company_name}" بنجاح')
        return redirect('stock:laboratory')
    
    return redirect('stock:laboratory')


@login_required
def laboratory_edit(request, pk):
    laboratory = get_object_or_404(Laboratory, pk=pk)
    
    if request.method == 'POST':
        laboratory.company_name = request.POST.get('company_name')
        laboratory.phone = request.POST.get('phone')
        laboratory.email = request.POST.get('email', '')
        laboratory.address = request.POST.get('address', '')
        laboratory.representative_name = request.POST.get('representative_name', '')
        laboratory.representative_phone = request.POST.get('representative_phone', '')
        laboratory.notes = request.POST.get('notes', '')
        laboratory.save()
        
        messages.success(request, f'تم تعديل المعمل "{laboratory.company_name}" بنجاح')
        return redirect('stock:laboratory')
    
    return redirect('stock:laboratory')


@login_required
def laboratory_delete(request, pk):
    laboratory = get_object_or_404(Laboratory, pk=pk)
    name = laboratory.company_name
    laboratory.delete()
    
    messages.success(request, f'تم حذف المعمل "{name}" بنجاح')
    return redirect('stock:laboratory')

# التعديلات المطلوبة على views.py

# استبدل دالة laboratory_detail_json بهذه النسخة المُصححة:

@login_required
def laboratory_detail_json(request, pk):
    """إرجاع تفاصيل المعمل كـ JSON"""
    laboratory = get_object_or_404(Laboratory, pk=pk)
    
    data = {
        'id': laboratory.id,
        'company_name': laboratory.company_name,
        'laboratory_code': laboratory.laboratory_code,
        'phone': laboratory.phone,
        'email': laboratory.email or '',
        'address': laboratory.address or '',
        'representative_name': laboratory.representative_name or '',
        'representative_phone': laboratory.representative_phone or '',
        'notes': laboratory.notes or '',
        'is_active': laboratory.is_active,
        # تم إزالة products_count لأن المعمل لا يحتوي على علاقة products
        'created_at': laboratory.created_at.strftime('%Y-%m-%d %H:%M'),
        'updated_at': laboratory.updated_at.strftime('%Y-%m-%d %H:%M'),
    }
    
    return JsonResponse(data)



# في ملف stock/views.py
# استبدل دالة products_list بهذه النسخة المحدثة:

@login_required
def products_list(request):
    """قائمة المنتجات"""
    
    search_query = request.GET.get('search', '')
    category_id = request.GET.get('category', '')
    status = request.GET.get('status', '')
    
    products = Product.objects.select_related('category', 'supplier').filter(is_active=True)
    
    if search_query:
        products = products.filter(
            Q(item_name__icontains=search_query) |
            Q(barcode__icontains=search_query)
        )
    
    if category_id:
        products = products.filter(category_id=category_id)
    
    if status == 'low':
        products = products.filter(quantity__lte=F('min_quantity'), quantity__gt=0)
    elif status == 'out':
        products = products.filter(quantity=0)
    
    products = products.order_by('-created_at')
    
    categories = Category.objects.filter(is_active=True)
    suppliers = Supplier.objects.filter(is_active=True)
    
    # Get logo URL from settings app (if available)
    logo_url = '/static/images/logo.png'  # Default logo
    try:
        from settings.models import GeneralSettings
        settings = GeneralSettings.objects.first()
        if settings and settings.logo:
            logo_url = settings.logo.url
    except:
        pass
    
    # Calculate stats
    total = products.count()
    low_stock = products.filter(quantity__lte=F('min_quantity'), quantity__gt=0).count()
    out_of_stock = products.filter(quantity=0).count()
    
    context = {
        'products': products,
        'categories': categories,
        'suppliers': suppliers,
        'search_query': search_query,
        'category_id': category_id,
        'status': status,
        'logo_url': logo_url,
        'stats': {
            'total': total,
            'low_stock': low_stock,
            'out_of_stock': out_of_stock,
        }
    }
    
    return render(request, 'stock/products_list.html', context)


# استبدل أيضاً دالة product_detail بهذه النسخة المحدثة:

@login_required
def product_detail(request, pk):
    """تفاصيل المنتج"""
    
    product = get_object_or_404(Product, pk=pk)
    movements = product.movements.all()[:20]
    
    # Get logo URL from settings app (if available)
    logo_url = '/static/images/logo.png'  # Default logo
    try:
        from settings.models import GeneralSettings
        settings = GeneralSettings.objects.first()
        if settings and settings.logo:
            logo_url = settings.logo.url
    except:
        pass
    
    context = {
        'product': product,
        'movements': movements,
        'logo_url': logo_url,
    }
    
    return render(request, 'stock/product_detail.html', context)

